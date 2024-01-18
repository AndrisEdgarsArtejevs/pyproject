import PyPDF2
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

data=[]

pdf_file1 = PyPDF2.PdfReader(open('student1.pdf', "rb"))
number_of_pages1 = len(pdf_file1.pages)
page1 = pdf_file1.pages[0]
lieciba = page1.extract_text()

p1 = lieciba.find("Angļu")
p01 = lieciba.find("mājasdarbi")
p2 = lieciba.find("Bioloģija")
p3 = lieciba.find("Datorika")
p4 = lieciba.find("tehnoloģijas")
p5 = lieciba.find("Fizika")
p6 = lieciba.find("Ģeogrāfija")
p7 = lieciba.find("Krievu")
p8 = lieciba.find("Ķīmija")
p9 = lieciba.find("Latviešu")
p10 = lieciba.find("vēstur e")
p11 = lieciba.find("Literatūra")
p12 = lieciba.find("Matemātika")
p13 = lieciba.find("Mūzika")
p14 = lieciba.find("zinības")
p15 = lieciba.find("Sports")
p16 = lieciba.find("māksla")
p17 = lieciba.find("Kavētās stundas")

def clean_data(subject_data):
    return subject_data.replace('(p.d.)',' ').replace('nv','').replace('\n', '').replace('n', '').replace('|','').replace(' ,',',').replace('%','% ' )

subjects = [
    ("English", p1 + 13, p01),
    ("Biology", p2 + 9, p3),
    ("Computing", p3 + 9, p4 - 11),
    ("Technologies", p4 + 12, p5),
    ("Physics", p5 + 6, p6),
    ("Geography", p6 + 11, p7),
    ("Russian", p7 + 14, p8),
    ("Chemistry", p8 + 9, p9),
    ("Latvian", p9 + 15, p10),
    ("History", p10 + 8, p11),
    ("Literature", p11 + 10, p12),
    ("Math", p12 + 10, p13),
    ("Music", p13 + 6, p14),
    ("Social Sciences", p14 + 7, p15),
    ("Sports", p15 + 18, p16),
    ("Visual Arts", p16 + 6, p17),
]

for name, start, end in subjects:
    subject_data = clean_data(lieciba[start:end])
    filtered_data = [int(i) for i in subject_data.split() if i.strip().isdigit()]
    average = sum(filtered_data) / len(filtered_data)
    # print([name, filtered_data, round(average, 2)])
    data.append([name, round(average)])
data.insert(0, ['Priekšmeti', '1. sem. grades'])

print(data)

wb = Workbook()
ws = wb.active

for row_num, (label, value) in enumerate(data, start=1):
    ws.cell(row=row_num, column=1, value=label)
    ws.cell(row=row_num, column=2, value=value)

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Analyse of grades"
chart.x_axis_title = "Category"
chart.y_axis_title = "Value"

data_range = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(data))
labels_range = Reference(ws, min_col=1, min_row=2, max_row=len(data))

chart.add_data(data_range, titles_from_data=True)
chart.set_categories(labels_range)
chart.width = 15 
chart.height = 10  

data_labels = DataLabelList()
data_labels.showVal = True

chart.legend.position = 'b'  # Set the legend position to bottom

chart.dataLabels = data_labels
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
chart.gapWidth = 47
ws.add_chart(chart, "A19")
chart.y_axis.scaling.max = 10

wb.save("chart.xlsx")